from django.db import models
import os
from django.conf import settings
from zipfile import ZipFile
from openpyxl import load_workbook
from uuid import uuid4


class FileInfo(models.Model):
    FILE_TYPE_CHOICES = (
        ('stopwords_list', 'stopwords_list'),
        ('symbols_list', 'symbols_list'),
        ('news', 'news'),
        ('documents', 'documents'),
    )
    ftype = models.CharField(
        max_length=32,
        null=False,
        blank=False,
        choices=FILE_TYPE_CHOICES,
    )
    fname = models.CharField(
        max_length=512,
        null=False,
        blank=False,
    )
    FILE_FORMAT_CHOICES = (
        ('xlsx', 'xlsx'),
        ('csv', 'csv'),
        ('json', 'json'),
        ('pkl', 'pkl'),
    )
    fformat = models.CharField(
        max_length=8,
        null=False,
        blank=False,
        choices=FILE_FORMAT_CHOICES,
    )
    fpath = models.FilePathField(
        null=False,
        blank=False,
    )
    _load_complate = models.BooleanField(default=False)

    def load_complate(self, from_which_row=None, up_to_which_row=None):
        if not self._load_complate:
            return self._load_complate
        if ((from_which_row == self.from_which_row) or (from_which_row is None)) and (
                (up_to_which_row == self.up_to_which_row) or up_to_which_row is None):
            return self._load_complate
        self._load_complate = False
        return self._load_complate

    from_which_row = models.PositiveIntegerField(default=0)
    up_to_which_row = models.PositiveIntegerField(default=0)


class File:
    def __init__(self, file_name):
        self.files_path = os.path.join(settings.MEDIA_ROOT, 'files')
        if not os.path.exists(self.files_path):
            os.mkdir(self.files_path)
        self.info = FileInfo.objects.filter(fname=file_name).last()

    def is_complate(self, from_which_row=None, up_to_which_row=None):
        if self.info is None:
            return False
        return self.info.load_complate(from_which_row=from_which_row, up_to_which_row=up_to_which_row)

    def save(self, file_path=None, file_type='news', complate=None, from_which_row=None, up_to_which_row=None):
        if file_path is not None:
            _uuid = uuid4()
            zpath = os.path.join(self.files_path, f'{_uuid}.zip')
            fname = os.path.basename(file_path)
            with ZipFile(zpath, 'w') as zfile:
                zfile.write(file_path, fname)

            file_info = FileInfo()
            file_info.ftype = file_type
            file_info.fname = fname
            fformat = fname.rsplit('.')[-1]
            file_info.fformat = fformat

            file_info.fpath = zpath
            file_info.save()
            self.info = file_info
            return True
        if complate is not None:
            if from_which_row is not None:
                self.info.from_which_row = from_which_row
            if up_to_which_row is not None:
                self.info.up_to_which_row = up_to_which_row
            self.info._load_complate = complate
            self.info.save()
            return True
        return False

    def load(self, to_be_continued=True, from_which_row=0, up_to_which_row=None):
        if self.info is None:
            return []

        tmp_dir = os.path.join(settings.BASE_DIR, 'tmp')
        if not os.path.exists(tmp_dir):
            os.mkdir(tmp_dir)

        with ZipFile(self.info.fpath, 'r') as zfile:
            zfile.extractall(tmp_dir)

        file_path = os.path.join(tmp_dir, self.info.fname)

        if self.info.fformat == 'xlsx':
            file = load_workbook(file_path, read_only=True)
            afile = file.active
            if to_be_continued:
                # if self.info.load_complate:
                #     return []
                from_which_row = self.info.which_row
            _data = []
            first = True
            header = []
            for i, row in enumerate(afile.iter_rows(values_only=True, max_row=up_to_which_row)):
                if first:
                    header = row
                    first = False
                    continue
                if row[2] == 'آموزشي' or row[2] == 'ورزشي' or row[2] == 'مذهبي' or row[2] == 'سياسي':
                    print(row[2])
                    if i < from_which_row:
                        continue
                    _data.append(row)
            file.close()
            os.remove(file_path)
            return _data, header
