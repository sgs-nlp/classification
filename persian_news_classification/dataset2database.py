import json
import logging
import openpyxl
import pickle
import os

from nvd.pre_processing import normilizer, tokenizer, without_stopword

from .models import Reference, add_reference, add_category, add_news, add_stopword, add_word, File, update_reference


def add2database(file_name: str) -> int:
    reference_obj = Reference.objects.filter(title=file_name).first()
    if reference_obj:
        if reference_obj.load_complate_flag:
            logging.info('The desired datset is available in the database.')
            return reference_obj
        else:
            if reference_obj.stopwords_list is None:
                logging.info('Started storing persian symbols in the database.')
                if not os.path.isfile('uploads/PERSIAN_SYMBOLS.pkl'):
                    from nvd.symbols import LIST as PERSIAN_SYMBOLS
                    logging.info('Started persian symbols pkl file craeted.')
                    with open('uploads/PERSIAN_SYMBOLS.pkl', 'ab') as pkl_file:
                        pickle.dump(PERSIAN_SYMBOLS, pkl_file)
                    logging.info('create persian symbols pkl file complate.')
                    for word in PERSIAN_SYMBOLS:
                        add_word(word)
                logging.info('Persian symbols storage in the database is complete.')
                # reference_id = add_reference(file_name)
                logging.info('Started stopwords list loading.')
                if not os.path.isfile('uploads/stopwords_list.pkl'):
                    stop_word_list_name = f'{file_name}.persian.stopword.json'
                    with open(stop_word_list_name, 'r', encoding='utf-8') as file:
                        stopwords_list = file.read()
                        stopwords_list = json.loads(stopwords_list)
                    logging.info('Stopwords list loaded.')
                    logging.info('Started storing persian stopwords list in the database.')
                    for word in stopwords_list:
                        add_stopword(word, reference_obj.pk)
                    logging.info('Started stopwords pkl file craeted.')
                    with open('uploads/stopwords_list.pkl', 'ab') as pkl_file:
                        pickle.dump(stopwords_list, pkl_file)
                    logging.info('create stopwords pkl file complate.')
                    update_reference(reference_id=reference_obj.pk, stopwords_list=stopwords_list)
                logging.info('Persian stopwords storage in the database is complete.')
            logging.info('Started news file loading.')
            wb_obj = openpyxl.load_workbook(file_name)
            logging.info('News file loaded.')
            sheet = wb_obj.active
            first = True
            column_title_list = []
            row_number = 1
            category_dict = {}
            logging.info('Started storing persian news in the database.')
            for row in sheet.iter_rows(max_row=50):
                col = []
                for cell in row:
                    col.append(cell.value)
                if first:
                    column_title_list = col
                    first = False
                    continue
                _data = {}
                for i in range(len(column_title_list)):
                    _data[column_title_list[i]] = col[i]

                logging.info(f'Started storing category title in the database.')
                if _data['Category'] not in category_dict:
                    category_id = add_category(_data['Category'], reference_obj.pk)
                    category_dict[_data['Category']] = category_id
                category_id = category_dict[_data['Category']]
                logging.info(f'Category title storage in the database is complete.')
                logging.info(f'Started text normalisering')
                titr_string = normilizer(_data['Titr'])
                content_string = normilizer(_data['Content'])
                logging.info(f'Text normalisering complate.')
                logging.info(f'Started storing the number {row_number} news item in the database.')
                add_news(
                    titr_string=titr_string,
                    content_string=content_string,
                    category_id=category_id,
                    reference_id=reference_obj.pk,
                )
                logging.info(f'The number {row_number} news item storage in the database is complete.')
                row_number += 1
            update_reference(
                reference_id=reference_obj.pk,
                titr_string_flag=True,
                titr_string_code_flag=True,
                titr_words_flag=True,
                titr_words_code_flag=True,
                titr_words_without_stopword_flag=True,
                titr_words_without_stopword_code_flag=True,
                content_string_flag=True,
                content_string_code_flag=True,
                content_words_flag=True,
                content_words_code_flag=True,
                content_words_without_stopword_flag=True,
                content_words_without_stopword_code_flag=True,
                category_flag=True,
                length=row_number,
                keywords_flag=True,
            )
            logging.info('Started categories pkl file craeted.')
            if not os.path.isfile('uploads/categories.pkl'):
                from .models import Category
                _categories = Category.objects.filter(reference_id=reference_obj.pk).all()
                categories = []
                for cat in _categories:
                    categories.append(dict(cat))
                with open('uploads/categories.pkl', 'ab') as pkl_file:
                    pickle.dump(categories, pkl_file)
                update_reference(reference_id=reference_obj.pk, categories_list=categories)
            logging.info('create categories pkl file complate.')
            update_reference(reference_id=reference_obj.pk, load_complate_flag=True)
            return reference_obj.pk
    else:
        logging.info('Started storing persian symbols in the database.')
        if not os.path.isfile('uploads/PERSIAN_SYMBOLS.pkl'):
            from nvd.symbols import LIST as PERSIAN_SYMBOLS
            logging.info('Started persian symbols pkl file craeted.')
            with open('uploads/PERSIAN_SYMBOLS.pkl', 'ab') as pkl_file:
                pickle.dump(PERSIAN_SYMBOLS, pkl_file)
            logging.info('create persian symbols pkl file complate.')
            for word in PERSIAN_SYMBOLS:
                add_word(word)
        logging.info('Persian symbols storage in the database is complete.')
        reference_id = add_reference(file_name)
        logging.info('Started stopwords list loading.')
        if not os.path.isfile('uploads/stopwords_list.pkl'):
            stop_word_list_name = f'{file_name}.persian.stopword.json'
            with open(stop_word_list_name, 'r', encoding='utf-8') as file:
                stopwords_list = file.read()
                stopwords_list = json.loads(stopwords_list)
            logging.info('Stopwords list loaded.')
            logging.info('Started storing persian stopwords list in the database.')
            for word in stopwords_list:
                add_stopword(word, reference_id)
            logging.info('Started stopwords pkl file craeted.')
            with open('uploads/stopwords_list.pkl', 'ab') as pkl_file:
                pickle.dump(stopwords_list, pkl_file)
            logging.info('create stopwords pkl file complate.')
            update_reference(reference_id=reference_id, stopwords_list=stopwords_list)
        logging.info('Persian stopwords storage in the database is complete.')
        logging.info('Started news file loading.')
        wb_obj = openpyxl.load_workbook(file_name)
        logging.info('News file loaded.')
        sheet = wb_obj.active
        first = True
        column_title_list = []
        row_number = 1
        category_dict = {}
        logging.info('Started storing persian news in the database.')
        for row in sheet.iter_rows(max_row=50):
            col = []
            for cell in row:
                col.append(cell.value)
            if first:
                column_title_list = col
                first = False
                continue
            _data = {}
            for i in range(len(column_title_list)):
                _data[column_title_list[i]] = col[i]

            logging.info(f'Started storing category title in the database.')
            if _data['Category'] not in category_dict:
                category_id = add_category(_data['Category'], reference_id)
                category_dict[_data['Category']] = category_id
            category_id = category_dict[_data['Category']]
            logging.info(f'Category title storage in the database is complete.')

            logging.info(f'Started text normalisering')
            titr_string = normilizer(_data['Titr'])
            content_string = normilizer(_data['Content'])
            logging.info(f'Text normalisering complate.')

            logging.info(f'Started storing the number {row_number} news item in the database.')
            add_news(
                titr_string=titr_string,
                content_string=content_string,
                category_id=category_id,
                reference_id=reference_id,
            )
            logging.info(f'The number {row_number} news item storage in the database is complete.')
            row_number += 1
        update_reference(
            reference_id=reference_id,
            titr_string_flag=True,
            titr_string_code_flag=True,
            titr_words_flag=True,
            titr_words_code_flag=True,
            titr_words_without_stopword_flag=True,
            titr_words_without_stopword_code_flag=True,
            content_string_flag=True,
            content_string_code_flag=True,
            content_words_flag=True,
            content_words_code_flag=True,
            content_words_without_stopword_flag=True,
            content_words_without_stopword_code_flag=True,
            category_flag=True,
            length=row_number,
            keywords_flag=True,
        )
        logging.info('Started categories pkl file craeted.')
        if not os.path.isfile('uploads/categories.pkl'):
            from .models import Category
            _categories = Category.objects.filter(reference_id=reference_id).all()
            categories = []
            for cat in _categories:
                categories.append(dict(cat))
            with open('uploads/categories.pkl', 'ab') as pkl_file:
                pickle.dump(categories, pkl_file)
                update_reference(reference_id=reference_id, categories_list=categories)
        logging.info('create categories pkl file complate.')
        update_reference(reference_id=reference_id, load_complate_flag=True)
        return reference_id
